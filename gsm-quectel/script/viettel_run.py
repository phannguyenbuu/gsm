#finally 25/04/2024
import os
import re
import time
import json 
import tkinter as tk
from tkinter import ttk
from datetime import datetime
from openpyxl import Workbook
import pytz
from selenium import webdriver
from tkinter import messagebox
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from selenium.webdriver.common.by import By
from cryptography.fernet import Fernet
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import sys

stop_flag = False
dir_ico = "tools\\hpk\\hpk.ico"
driver_link = "https://kpp.bankplus.vn"
folder_result = "ket_qua"
title = "Thông báo"
glo_payment_inter = "payment_internet"
glo_payment_card = "deb_cart"
glo_lookup_card = "lookup_cart"
glo_lookup_ftth = "lookup_ftth"
glo_payment_evn = "deb_evn"
dir_config = "tools\\hpk\\config.json"
uri = "mongodb+srv://huytq0104:huypro7690@hpk.jdkmqqs.mongodb.net/?retryWrites=true&w=majority&appName=HPK"
copy_right = b"h_ThisAAutoToolVjppro-CopyRight-ByCAOAC7690="
stt_complete = "Đã xử lý"
stt_incomplete = "Chưa xử lý"

# Xuất file kết quả tra cứu dạng excel
def export_excel(data, name_dir):
    today = datetime.now().strftime("%H%M-%d-%m-%Y")
    try:
        export_dir = os.path.join(os.getcwd(), f"{folder_result}\\{name_dir}")
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        file_name = f"{today}.xlsx"
        file_path = os.path.join(export_dir, file_name)

        wb = Workbook()
        ws = wb.active
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        ws['A1'] = 'STT'
        ws['B1'] = 'Số thuê bao'
        ws['C1'] = 'Số tiền'
        ws['D1'] = 'Ghi chú'

        bold_font = Font(bold=True)
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=4):
            for cell in col:
                cell.font = bold_font
                cell.fill = yellow_fill
        for idx, (phone, amount, note) in enumerate(data, start=2):
            ws[f'A{idx}'] = idx - 1
            ws[f'B{idx}'] = phone
            ws[f'C{idx}'] = amount
            ws[f'D{idx}'] = note
        wb.save(file_path)

    except Exception as e:
        messagebox.showerror(title , f"*er09 - {e}")

def check_username():
    try:
        dl_info = WebDriverWait(driver, 4).until(EC.presence_of_element_located((By.CLASS_NAME, "dl-info-detail")))
        first_div = dl_info.find_element(By.XPATH, "./div[1]")
        span_ele = first_div.find_element(By.XPATH, ".//span")
        username = span_ele.text
        if dbfiles.get("username").strip() != username:
            messagebox.showerror(title, "Vui lòng sử dụng đúng tài khoản đã đăng ký")
            return False
        else:
            return True
    except Exception as e:
        messagebox.showerror(title, "Không tìm thấy tên tài khoản trên Viettel Pay Pro")
        return False

def amount_by_cbil(cbil, element, lookup):
    try:
        amount = "Không tìm thấy mã thuê bao"
        payment_id = None      
        html_content = element.get_attribute('outerHTML')
        soup = BeautifulSoup(html_content, 'html.parser')
        pay_content_groups = soup.find_all("div", class_="row pay-content mb-3")
        for group in pay_content_groups:
            p_tags = group.find_all("p")
            is_find = False
            for p_tag in p_tags:
                if cbil in p_tag.text:
                    is_find = True
                    break

            if lookup and is_find:
                button_tag = group.find("button", {"id": re.compile(r'payMoneyForm:btnView\d*')})
                if button_tag:
                    payment_id = button_tag['id']

            if is_find:
                for p_tag in p_tags:
                    if "VND" in p_tag.text:
                        str_price = p_tag.text.split("VND")[0].strip()
                        amount = int(str_price.replace(",", ""))
                        if amount >= 5000:
                            return True, amount, payment_id
                        else:
                            return False, amount, payment_id

        return False, amount, payment_id
    except Exception as e:
        return False, "Lỗi thanh toán", ""

# ------------------------------------------------------------------- Hàm xử lý thanh toán internet - tv
def payment_internet(tkinp_ctm, tkinp_ctmed, tkinp_pin, lbl_times):
    try:
        type_services = glo_payment_inter
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils, pin]):
            return False
        data = []
        for cbil in cbils:
            root.update()
            time.sleep(0.5)
            cbil = cbil.strip()
            times = int(times_exits[type_services])
            if times < 1:
                messagebox.showwarning(title, "Số lần sử dụng trong tháng đã hết")
                break
            elif not stop_flag and cbil.strip() != "":
                try:
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                    customer.clear()
                    customer.send_keys(cbil)
                    time.sleep(0.5)
                    payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located(( By.ID, "payMoneyForm:btnPay0")))
                    payment_button.click()
                    time.sleep(1)
                    WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                    element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                    is_amount, amount, payment_id = amount_by_cbil(cbil, element41 ,True)
                    if not is_amount:
                        data.append([cbil, amount, stt_complete])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        continue
                    else:
                        payment_btn1 = WebDriverWait(driver, 16).until(EC.presence_of_element_located((By.ID, payment_id)))
                        payment_btn1.click()
                        pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:pinId")))
                        pin_id.clear()
                        pin_id.send_keys(pin)
                        pay_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay")))
                        pay_btn.click()
                        try:
                            cfm_modal = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "payMoneyForm:dlgConfirm_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                        except:
                            pass
                        confirm_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:yesId0")))
                        confirm_btn.click()
                        data.append([cbil, amount, stt_complete])
                        update_file_config(type_services, lbl_times)
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                       
                except Exception as e:
                    data.append([cbil, 0, stt_incomplete])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                # break
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Thanh toán TV - Internet"
            export_excel(data, name_dir)
    except Exception as e:
        messagebox.showerror(title , f"*er16 - {e}")

#-------------------------------------------------------------------- kết thúc thanh toán internet

# ----------------------------------------------------- Hàm xử lý tra cứu ftth

def lookup_ftth(tkinp_ctm, tkinp_ctmed, lbl_times):
    try:
        type_services = glo_lookup_ftth
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils]):
            return False
        data = []
        for cbil in cbils:
            root.update()
            time.sleep(1)
            cbil = cbil.strip()
            times = int(times_exits[type_services])
            if times < 1:
                messagebox.showwarning(title, "Số lần sử dụng trong tháng đã hết")
                break
            elif not stop_flag and cbil.strip() != "":
                try:
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:contractCode")))
                    customer.clear()
                    customer.send_keys(cbil)
                    payment_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located(( By.ID, "payMoneyForm:btnPay0")))
                    payment_button.click()
                    time.sleep(1)
                    WebDriverWait(driver, 16).until(EC.invisibility_of_element_located((By.ID, "payMoneyForm:j_idt6_modal")))
                    element41 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt41")))
                    is_amount, amount, payment_id = amount_by_cbil(cbil, element41 ,False)
                    data.append([cbil, amount, stt_complete])
                    update_file_config(type_services, lbl_times)
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    continue
                except Exception as e:
                    data.append([cbil, 0, stt_incomplete])
                    insert_ctmed(tkinp_ctmed, f"{cbil} - Lỗi")
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    continue
                # break
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Tra cứu FTTH"
            export_excel(data, name_dir)
    except Exception as e:
        messagebox.showerror(title , f"*er17 - {e}")

#------------------------------------------------------ kết thúc tra cứu ftth

# -----------------------------------------------------------------Nạp tiền đa mạng

def payment_phone(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkcbb_form, tkcbb_amount, lbl_times):
    try:
        type_services = glo_payment_card
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        pin = tkinp_pin.get()
        cbb_type = tkcbb_form.get()
        type_sub = handle_choose_select(cbb_type)
        if type_sub == 1:
            amount = tkcbb_amount.get()
            isnext = valid_data([cbils, pin, amount])
            if isnext:
                rsl_amount = handle_choose_amount(amount)
        else:
            isnext = valid_data([cbils, pin])

        if not isnext:
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            times = int(times_exits[type_services])
            if times < 1:
                messagebox.showwarning(title, "Số lần sử dụng trong tháng đã hết")
                break
            elif not stop_flag and cbil.strip() != "":
                driver.refresh()
                time.sleep(2)
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(1)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                time.sleep(0.5)
                if type_sub == 1:
                    try:
                        try:
                            cfm_modalTT = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                            time.sleep(1)
                        except:
                            pass
                        spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                        spl_lbl.click()
                        spl_0 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_0")))
                        spl_0.click()
                        cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                        cfm_pay.click()
                    except:
                        pass
                    script = f"""
                    var element = document.querySelector('input[id="indexForm:subAmountId:{rsl_amount}"]').closest('div');
                    if (!element.classList.contains('ui-state-active')) {{
                        element.click();
                    }}
                    """
                    driver.execute_script(script)
                else:
                    try:
                        try:
                            cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                        except:
                            pass
                        try:
                            time.sleep(0.5)
                            spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                            spl_lbl.click()
                            time.sleep(0.5)
                            spl_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_1")))
                            spl_1.click()
                            time.sleep(0.5)
                            cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                            cfm_pay.click()
                            time.sleep(0.5)
                        except:
                            pass
                        try:
                            btn_check = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnCheck")))
                            btn_check.click()
                        except:
                            pass
                        lbl_debt = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:debtId_input")))
                        debt_str = lbl_debt.get_attribute('value')
                        debt = int(debt_str.replace(".","").replace(",",""))
                        if debt >= 5000:
                            inp_amount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:transAmountId_input")))
                            inp_amount.clear()
                            inp_amount.send_keys(debt)
                        else:
                            data.append([cbil, debt, stt_complete])
                            tkinp_ctm.delete("1.0", "1.end+1c")
                            insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                            continue
                    except:
                        data.append([cbil, 0, stt_complete])
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - không nợ cước")
                        continue
                try:
                    pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:pinId")))
                    pin_id.clear()
                    pin_id.send_keys(pin)
                    btn_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnPay")))
                    btn_pay.click()
                    try:
                        cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirm_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                    except:
                        pass
                    time.sleep(0.5)
                    btn_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesIdCard")))
                    btn_confirm.click()
                    if type_sub == 1:
                        data.append([cbil, amount, stt_complete])
                        update_file_config(type_services, lbl_times)
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    else:
                        data.append([cbil, debt, stt_complete])
                        update_file_config(type_services, lbl_times)
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                except:
                    data.append([cbil, 0, stt_incomplete])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Nạp tiền đa mạng"
            export_excel(data, name_dir)
    except Exception as e:
        messagebox.showerror(title , f"*er18 - {e}")

# -----------------------------------------------------------------Kết thúc nạp tiền đa mạng

# --------------------------------------------------------Hàm tra cứu tiền đa mạng

def lookup_card(tkinp_ctm, tkinp_ctmed, lbl_times):
    try:
        type_services = glo_payment_card
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils]):
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            times = int(times_exits[type_services])
            if times < 1:
                messagebox.showwarning(title, "Số lần sử dụng trong tháng đã hết")
                break
            elif not stop_flag and cbil.strip() != "":
                driver.refresh()
                time.sleep(2)
                try:
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                    time.sleep(1)
                except:
                    time.sleep(2)
                    phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                    phonenum.clear()
                    phonenum.send_keys(cbil)
                    phonenum.send_keys(Keys.TAB)
                
                try:
                    try:
                        cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                    except:
                        pass
                    try:
                        time.sleep(0.5)
                        spl_lbl = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                        spl_lbl.click()
                        time.sleep(0.5)
                        spl_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_1")))
                        spl_1.click()
                        time.sleep(0.5)
                        cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                        cfm_pay.click()
                        time.sleep(0.5)
                    except:
                        pass
                    try:
                        btn_check = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnCheck")))
                        btn_check.click()
                    except:
                        pass
                    lbl_debt = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:debtId_input")))
                    debt_str = lbl_debt.get_attribute('value')
                    debt = int(debt_str.replace(".","").replace(",",""))
                    
                    data.append([cbil, debt, stt_complete])
                    update_file_config(type_services, lbl_times)
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {debt}")
                    continue
                except:
                    data.append([cbil, "Không tìm thấy nợ cước", stt_incomplete])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - null")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Tra nợ trả sau"
            export_excel(data, name_dir)
    except Exception as e:
        messagebox.showerror(title , f"*er19 - {e}")

# --------------------------------------------------------Kết thúc tra cứu tiền đa mạng

# --------------------------------------------------------------------------Hàm nạp tiền mạng Viettel
def payment_viettel(tkinp_ctm, tkinp_ctmed, tkinp_pin, tkinp_amount, lbl_times):
        try:
            type_services = glo_payment_card
            delete_ctmed(tkinp_ctmed)
            update_stop_flag()
            pin = tkinp_pin.get()
            amount = tkinp_amount.get()
            cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
            if not valid_data([cbils, pin, amount]):
                return False
            data = []
            for cbil in cbils:
                root.update()
                time.sleep(0.5)
                cbil = cbil.strip()
                times = int(times_exits[type_services])
                if times < 1:
                    messagebox.showwarning(title, "Số lần sử dụng trong tháng đã hết")
                    break
                elif not stop_flag and cbil.strip() != "":
                    time.sleep(0.5)
                    try:
                        phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                        phonenum.clear()
                        phonenum.send_keys(cbil)
                        phonenum.send_keys(Keys.TAB)
                        time.sleep(0.5)
                    except:
                        time.sleep(2)
                        phonenum = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:phoneNumberId")))
                        phonenum.clear()
                        phonenum.send_keys(cbil)
                        phonenum.send_keys(Keys.TAB)
                    try:
                        cfm_modalTT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirmTT_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modalTT)
                        time.sleep(0.5)
                        spl_lbl = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "indexForm:supplier")))
                        spl_lbl.click()
                        time.sleep(0.5)
                        spl_0 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:supplier_0")))
                        spl_0.click()
                        time.sleep(0.5)
                        cfm_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesTTId")))
                        cfm_pay.click()
                        time.sleep(0.5)
                    except:
                        pass
                    try:
                        inp_amount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:transAmountId_input")))
                        inp_amount.clear()
                        inp_amount.send_keys(amount)
                        time.sleep(0.5)
                        pin_id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:pinId")))
                        pin_id.clear()
                        pin_id.send_keys(pin)
                        time.sleep(0.5)
                        btn_pay = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:btnPay")))
                        btn_pay.click()
                        time.sleep(0.5)
                        try:
                            cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "indexForm:dlgConfirm_modal")))
                            driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                        except:
                            pass
                        btn_confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "indexForm:yesIdCard")))
                        btn_confirm.click()
                        data.append([cbil, amount, stt_complete])
                        update_file_config(type_services, lbl_times)
                        insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                    except:
                        data.append([cbil, 0, stt_incomplete])
                        insert_ctmed(tkinp_ctmed, f"{cbil} - null")
                        tkinp_ctm.delete("1.0", "1.end+1c")
                        continue
            time.sleep(2)
            if len(data) > 0:
                name_dir = "Nạp tiền mạng Viettel"
                export_excel(data, name_dir)
        except Exception as e:
            messagebox.showerror(title , f"*er20 - {e}")

# --------------------------------------------------------------------------Kết thúc nạp tiền mạng Viettel

# ---------------------------------------------------- Hàm xử lý gạch điện EVN
def debt_electric(tkinp_ctm, tkinp_ctmed, tkinp_phone, tkinp_pin, lbl_times):
    try:
        type_services = glo_payment_evn
        delete_ctmed(tkinp_ctmed)
        update_stop_flag()
        pin = tkinp_pin.get()
        phone = tkinp_phone.get()
        cbils = tkinp_ctm.get("1.0", "end-1c").splitlines()
        if not valid_data([cbils, pin, phone]):
            return False
        data = []
        for cbil in cbils:
            cbil = cbil.strip()
            root.update()
            time.sleep(1)
            times = int(times_exits[type_services])
            if times < 1:
                messagebox.showwarning(title, "Số lần sử dụng trong tháng đã hết")
                break
            elif not stop_flag and cbil.strip() != "":
                try:
                    customer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:billCodeId")))
                    customer.clear()
                    customer.send_keys(cbil)
                    time.sleep(0.5)
                    phonenumber = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:benMsisdnId")))
                    phonenumber.clear()
                    phonenumber.send_keys(phone)
                    time.sleep(0.5)
                    pinid = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:pinId")))
                    pinid.clear()
                    pinid.send_keys(pin)
                    time.sleep(0.5)
                    payment = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:btnPay")))
                    payment.click()
                    time.sleep(0.5)
                    try:
                        cfm_modal = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "payMoneyForm:dlgConfirm_modal")))
                        driver.execute_script("arguments[0].style.zIndex = '-99';", cfm_modal)
                    except:
                        pass
                    lblamount = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:j_idt49")))
                    try:
                        text_of_amount = lblamount.text
                        amount_str = text_of_amount.replace('VND', '').replace('.', '')
                        amount = int(amount_str)
                    except:
                        amount = lblamount.text
                    time.sleep(0.5)
                    confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "payMoneyForm:yesIdEVN")))
                    confirm.click()
                    data.append([cbil, amount, stt_complete])
                    update_file_config(type_services, lbl_times)
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - {amount}")
                    continue
                except Exception as e:
                    data.append([cbil, amount, stt_incomplete])
                    tkinp_ctm.delete("1.0", "1.end+1c")
                    insert_ctmed(tkinp_ctmed, f"{cbil} - null")
                    continue
        time.sleep(2)
        if len(data) > 0:
            name_dir = "Thanh toán điện EVN"
            export_excel(data, name_dir)
    except:
        messagebox.showerror(title , f"*er22 - {e}")

def handle_choose_services(event, choose, main_frm):
    element = choose.get()
    if element == "Thanh toán TV - Internet":
        clear_widgets(main_frm)
        form_payment_internet()
    elif element == "Nạp tiền đa mạng":
        clear_widgets(main_frm)
        form_payment_phone()
    elif element == "Tra cứu nợ thuê bao trả sau":
        clear_widgets(main_frm)
        form_lookup_card()
    elif element == "Tra cứu FTTH":
        clear_widgets(main_frm)
        form_lookup_ftth()
    elif element == "Nạp tiền mạng Viettel":
        clear_widgets(main_frm)
        form_payment_viettel()
    elif element == "Gạch điện EVN":
        clear_widgets(main_frm)
        form_debt_electric()
    root.update()
